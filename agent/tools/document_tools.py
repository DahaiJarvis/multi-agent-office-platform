"""文档处理原生工具

提供文档解析、摘要生成和文档对比功能。

工具列表：
  -------------------------------------------------------------------------
  native_document_parse: 解析 PDF/Word/TXT 文档
    - 延迟分层: fast
    - 权限级别: sensitive
    - 注册方式: 懒注册（依赖文件系统）

  native_document_summarize: 生成文档摘要
    - 延迟分层: slow
    - 权限级别: read_only
    - 注册方式: 懒注册（依赖 LLM）

  native_document_compare: 对比两份文档差异
    - 延迟分层: slow
    - 权限级别: read_only
    - 注册方式: 懒注册（依赖 LLM）
  -------------------------------------------------------------------------

安全约束：
  - native_document_parse 的 file_path 必须通过 path_validator.validate_file_path() 校验
  - permission_level=sensitive 的工具在安全护栏中需用户确认
"""

import json
import logging
from typing import Any

from autogen_core.tools import FunctionTool

from agent.tools.base import LatencyTier, NativeToolMeta, PermissionLevel
from agent.tools.registry import NativeToolRegistry

logger = logging.getLogger(__name__)


async def _document_parse(file_path: str, max_pages: int = 50) -> str:
    """解析文档内容

    支持 PDF、Word、TXT 等格式，返回文档文本内容。
    文件路径必须通过安全校验。

    Args:
        file_path: 文件路径，必须在白名单目录中
        max_pages: 最大解析页数，默认 50 页

    Returns:
        JSON 格式的文档内容
    """
    if not file_path or not file_path.strip():
        return json.dumps({"error": "文件路径不能为空", "content": "", "metadata": {}}, ensure_ascii=False)

    try:
        from agent.tools.path_validator import validate_file_path, PathValidationError

        safe_path = validate_file_path(file_path.strip())
    except PathValidationError as e:
        return json.dumps({"error": f"路径校验失败: {e.reason}", "content": "", "metadata": {}}, ensure_ascii=False)

    try:
        import os
        _, ext = os.path.splitext(safe_path)
        ext = ext.lower()

        content = ""
        metadata: dict[str, Any] = {
            "file_path": safe_path,
            "file_extension": ext,
            "file_size": 0,
        }

        if os.path.isfile(safe_path):
            metadata["file_size"] = os.path.getsize(safe_path)

        if ext == ".txt" or ext == ".md":
            with open(safe_path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()

        elif ext == ".pdf":
            content = await _parse_pdf(safe_path, max_pages)

        elif ext in (".docx", ".doc"):
            content = await _parse_docx(safe_path)

        elif ext == ".csv":
            content = await _parse_csv(safe_path)

        elif ext == ".xlsx":
            content = await _parse_xlsx(safe_path)

        else:
            return json.dumps({
                "error": f"不支持的文件格式: {ext}",
                "content": "",
                "metadata": metadata,
            }, ensure_ascii=False)

        metadata["content_length"] = len(content)
        metadata["pages_parsed"] = min(max_pages, metadata.get("pages_parsed", 1))

        result = {
            "file_path": safe_path,
            "content": content,
            "metadata": metadata,
        }
        return json.dumps(result, ensure_ascii=False)
    except Exception as e:
        logger.error("文档解析失败: file_path=%s, error=%s", file_path[:100], e)
        return json.dumps({"error": f"文档解析失败: {str(e)}", "content": "", "metadata": {}}, ensure_ascii=False)


async def _parse_pdf(file_path: str, max_pages: int) -> str:
    """解析 PDF 文件

    Args:
        file_path: PDF 文件路径
        max_pages: 最大解析页数

    Returns:
        提取的文本内容
    """
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(file_path)
        pages = []
        for i, page in enumerate(doc):
            if i >= max_pages:
                break
            pages.append(page.get_text())
        doc.close()
        return "\n\n".join(pages)
    except ImportError:
        try:
            from PyPDF2 import PdfReader

            reader = PdfReader(file_path)
            pages = []
            for i, page in enumerate(reader.pages):
                if i >= max_pages:
                    break
                text = page.extract_text()
                if text:
                    pages.append(text)
            return "\n\n".join(pages)
        except ImportError:
            logger.warning("未安装 PDF 解析库（PyMuPDF 或 PyPDF2），无法解析 PDF")
            return "[PDF 解析需要安装 PyMuPDF 或 PyPDF2 库]"
    except Exception as e:
        logger.error("PDF 解析异常: %s", e)
        return f"[PDF 解析失败: {e}]"


async def _parse_docx(file_path: str) -> str:
    """解析 Word 文件

    Args:
        file_path: Word 文件路径

    Returns:
        提取的文本内容
    """
    try:
        from docx import Document

        doc = Document(file_path)
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n".join(paragraphs)
    except ImportError:
        logger.warning("未安装 python-docx 库，无法解析 Word 文档")
        return "[Word 解析需要安装 python-docx 库]"
    except Exception as e:
        logger.error("Word 解析异常: %s", e)
        return f"[Word 解析失败: {e}]"


async def _parse_csv(file_path: str) -> str:
    """解析 CSV 文件

    Args:
        file_path: CSV 文件路径

    Returns:
        CSV 内容文本
    """
    import csv

    try:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            rows = list(reader)
        if not rows:
            return ""
        header = rows[0]
        content_lines = [", ".join(header)]
        for row in rows[1:]:
            content_lines.append(", ".join(row))
        return "\n".join(content_lines)
    except Exception as e:
        logger.error("CSV 解析异常: %s", e)
        return f"[CSV 解析失败: {e}]"


async def _parse_xlsx(file_path: str) -> str:
    """解析 Excel 文件

    Args:
        file_path: Excel 文件路径

    Returns:
        Excel 内容文本
    """
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        all_content = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_content.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(values_only=True):
                row_text = ", ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip(", "):
                    all_content.append(row_text)
        wb.close()
        return "\n".join(all_content)
    except ImportError:
        logger.warning("未安装 openpyxl 库，无法解析 Excel 文件")
        return "[Excel 解析需要安装 openpyxl 库]"
    except Exception as e:
        logger.error("Excel 解析异常: %s", e)
        return f"[Excel 解析失败: {e}]"


async def _document_summarize(file_path: str, mode: str = "brief") -> str:
    """生成文档摘要

    解析文档内容并使用 LLM 生成摘要。

    Args:
        file_path: 文件路径，必须在白名单目录中
        mode: 摘要模式，brief(简要) / detailed(详细)

    Returns:
        JSON 格式的文档摘要
    """
    if not file_path or not file_path.strip():
        return json.dumps({"error": "文件路径不能为空", "summary": ""}, ensure_ascii=False)

    if mode not in ("brief", "detailed"):
        return json.dumps({"error": f"不支持的摘要模式: {mode}，仅允许 brief 和 detailed", "summary": ""}, ensure_ascii=False)

    try:
        from agent.tools.path_validator import validate_file_path, PathValidationError

        safe_path = validate_file_path(file_path.strip())
    except PathValidationError as e:
        return json.dumps({"error": f"路径校验失败: {e.reason}", "summary": ""}, ensure_ascii=False)

    try:
        parse_result = await _document_parse(safe_path)
        parsed = json.loads(parse_result)

        if "error" in parsed and parsed.get("content", "") == "":
            return json.dumps({"error": parsed["error"], "summary": ""}, ensure_ascii=False)

        content = parsed.get("content", "")
        if not content.strip():
            return json.dumps({"file_path": safe_path, "summary": "文档内容为空", "mode": mode}, ensure_ascii=False)

        from agent.core.model_client import get_lightweight_client
        from autogen_core.models import UserMessage

        client = get_lightweight_client()

        max_content = 8000 if mode == "detailed" else 4000
        truncated = content[:max_content]
        if len(content) > max_content:
            truncated += f"\n\n[文档内容已截断，原文共 {len(content)} 字符]"

        if mode == "detailed":
            prompt = (
                "请详细总结以下文档内容，包括：\n"
                "1. 文档的主要主题和目的\n"
                "2. 关键要点和核心论据\n"
                "3. 重要数据和结论\n"
                "4. 需要特别关注的细节\n\n"
                f"文档内容：\n{truncated}"
            )
        else:
            prompt = f"请简要总结以下文档内容（200字以内）：\n\n{truncated}"

        response = await client.create(
            messages=[UserMessage(source="user", content=prompt)],
            extra_create_args={"temperature": 0.3, "max_tokens": 1000 if mode == "detailed" else 500},
        )

        summary = response.content if isinstance(response.content, str) else str(response.content)

        result = {
            "file_path": safe_path,
            "summary": summary.strip(),
            "mode": mode,
            "content_length": len(content),
        }
        return json.dumps(result, ensure_ascii=False)
    except Exception as e:
        logger.error("文档摘要生成失败: file_path=%s, error=%s", file_path[:100], e)
        return json.dumps({"error": f"文档摘要生成失败: {str(e)}", "summary": ""}, ensure_ascii=False)


async def _document_compare(file_path_a: str, file_path_b: str, aspect: str = "general") -> str:
    """对比两份文档差异

    解析两份文档并使用 LLM 对比差异。

    Args:
        file_path_a: 第一份文档路径
        file_path_b: 第二份文档路径
        aspect: 对比维度，general(综合) / content(内容) / structure(结构) / data(数据)

    Returns:
        JSON 格式的对比结果，包含差异列表
    """
    if not file_path_a or not file_path_a.strip():
        return json.dumps({"error": "文档A路径不能为空", "differences": []}, ensure_ascii=False)
    if not file_path_b or not file_path_b.strip():
        return json.dumps({"error": "文档B路径不能为空", "differences": []}, ensure_ascii=False)

    try:
        from agent.tools.path_validator import validate_file_path, PathValidationError

        safe_path_a = validate_file_path(file_path_a.strip())
        safe_path_b = validate_file_path(file_path_b.strip())
    except PathValidationError as e:
        return json.dumps({"error": f"路径校验失败: {e.reason}", "differences": []}, ensure_ascii=False)

    try:
        parse_a = await _document_parse(safe_path_a)
        parse_b = await _document_parse(safe_path_b)

        content_a = json.loads(parse_a).get("content", "")
        content_b = json.loads(parse_b).get("content", "")

        if not content_a.strip() or not content_b.strip():
            return json.dumps({
                "file_path_a": safe_path_a,
                "file_path_b": safe_path_b,
                "differences": [],
                "message": "其中一份文档内容为空，无法对比",
            }, ensure_ascii=False)

        from agent.core.model_client import get_lightweight_client
        from autogen_core.models import UserMessage

        client = get_lightweight_client()

        max_len = 4000
        text_a = content_a[:max_len]
        text_b = content_b[:max_len]

        aspect_instructions = {
            "general": "请从内容、结构、数据等多个维度综合对比",
            "content": "请重点对比两份文档的内容差异",
            "structure": "请重点对比两份文档的结构和格式差异",
            "data": "请重点对比两份文档中的数据和数值差异",
        }
        instruction = aspect_instructions.get(aspect, aspect_instructions["general"])

        prompt = (
            f"{instruction}以下两份文档的差异，列出具体的差异点：\n\n"
            f"【文档A】\n{text_a}\n\n"
            f"【文档B】\n{text_b}\n\n"
            "请按以下格式输出差异列表：\n"
            "- 差异点1: ...\n"
            "- 差异点2: ...\n"
        )

        response = await client.create(
            messages=[UserMessage(source="user", content=prompt)],
            extra_create_args={"temperature": 0.2, "max_tokens": 1500},
        )

        comparison = response.content if isinstance(response.content, str) else str(response.content)

        differences = []
        for line in comparison.strip().split("\n"):
            line = line.strip()
            if line.startswith("- ") or line.startswith("* "):
                differences.append(line[2:].strip())
            elif line and not line.startswith("#"):
                differences.append(line)

        result = {
            "file_path_a": safe_path_a,
            "file_path_b": safe_path_b,
            "aspect": aspect,
            "differences": differences,
            "comparison": comparison.strip(),
        }
        return json.dumps(result, ensure_ascii=False)
    except Exception as e:
        logger.error("文档对比失败: error=%s", e)
        return json.dumps({"error": f"文档对比失败: {str(e)}", "differences": []}, ensure_ascii=False)


_DOCUMENT_PARSE_META = NativeToolMeta(
    name="native_document_parse",
    display_name="文档解析",
    description="解析 PDF、Word、TXT、CSV、Excel 等格式的文档，提取文本内容。文件路径必须在白名单目录中，且扩展名必须被允许。",
    category="document",
    parameters={
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "文档文件路径，必须在白名单目录中",
            },
            "max_pages": {
                "type": "integer",
                "description": "最大解析页数（PDF），默认 50 页",
                "default": 50,
                "minimum": 1,
                "maximum": 200,
            },
        },
        "required": ["file_path"],
    },
    latency_tier=LatencyTier.FAST,
    permission_level=PermissionLevel.SENSITIVE,
    timeout_seconds=60,
    requires_llm=False,
    tags=["document", "parse", "pdf", "word"],
)

_DOCUMENT_SUMMARIZE_META = NativeToolMeta(
    name="native_document_summarize",
    display_name="文档摘要",
    description="使用 AI 生成文档摘要，支持简要和详细两种模式。文件路径必须在白名单目录中。",
    category="document",
    parameters={
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string",
                "description": "文档文件路径",
            },
            "mode": {
                "type": "string",
                "description": "摘要模式: brief(简要) / detailed(详细)",
                "enum": ["brief", "detailed"],
                "default": "brief",
            },
        },
        "required": ["file_path"],
    },
    latency_tier=LatencyTier.SLOW,
    permission_level=PermissionLevel.READ_ONLY,
    timeout_seconds=60,
    requires_llm=True,
    tags=["document", "summary", "llm"],
)

_DOCUMENT_COMPARE_META = NativeToolMeta(
    name="native_document_compare",
    display_name="文档对比",
    description="使用 AI 对比两份文档的差异，支持综合、内容、结构、数据等对比维度。返回差异列表。",
    category="document",
    parameters={
        "type": "object",
        "properties": {
            "file_path_a": {
                "type": "string",
                "description": "第一份文档路径",
            },
            "file_path_b": {
                "type": "string",
                "description": "第二份文档路径",
            },
            "aspect": {
                "type": "string",
                "description": "对比维度: general(综合) / content(内容) / structure(结构) / data(数据)",
                "enum": ["general", "content", "structure", "data"],
                "default": "general",
            },
        },
        "required": ["file_path_a", "file_path_b"],
    },
    latency_tier=LatencyTier.SLOW,
    permission_level=PermissionLevel.READ_ONLY,
    timeout_seconds=90,
    requires_llm=True,
    tags=["document", "compare", "llm"],
)


def register_all(registry: NativeToolRegistry) -> None:
    """注册所有文档处理工具

    native_document_parse 依赖文件系统，native_document_summarize 和
    native_document_compare 依赖 LLM，均使用懒注册。

    Args:
        registry: 工具注册中心实例
    """

    def _create_parse_tool() -> FunctionTool:
        return FunctionTool(
            func=_document_parse,
            name="native_document_parse",
            description=_DOCUMENT_PARSE_META.description,
        )

    def _create_summarize_tool() -> FunctionTool:
        return FunctionTool(
            func=_document_summarize,
            name="native_document_summarize",
            description=_DOCUMENT_SUMMARIZE_META.description,
        )

    def _create_compare_tool() -> FunctionTool:
        return FunctionTool(
            func=_document_compare,
            name="native_document_compare",
            description=_DOCUMENT_COMPARE_META.description,
        )

    registry.register_lazy("native_document_parse", _create_parse_tool, _DOCUMENT_PARSE_META)
    registry.register_lazy("native_document_summarize", _create_summarize_tool, _DOCUMENT_SUMMARIZE_META)
    registry.register_lazy("native_document_compare", _create_compare_tool, _DOCUMENT_COMPARE_META)

    logger.debug("文档处理工具注册完成: native_document_parse, native_document_summarize, native_document_compare(均懒注册)")
